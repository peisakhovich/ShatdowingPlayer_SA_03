from openpyxl import load_workbook

from session.session import Session
from session.session_excel import SessionExcel


def create_test_session() -> Session:
    """
    Создаёт небольшую тестовую Session.

    Реальный plan_session.json для unit-тестов
    не используется.
    """

    session = Session()

    session.load_data({
        "user_id": 2,
        "user_nickname": "pga",

        "set": {
            "set_id": 2,
            "user_id": 2,
            "user_nickname": "pga",
            "set_index": 2,
            "set_name": "Polski B2",
            "set_description": "Testowy zestaw",
            "set_active": True,
            "set_create_date": "2026-07-31T14:29:01",
            "set_update_date": "2026-09-04T12:13:14",
            "items_count": 2,
        },

        "items": [
            {
                "item_id": 38,
                "item_order": 1,
                "phrase_id": 57,
                "difficulty": 1,
                "phrase_text": "Co to jest napięcie elektryczne?",
                "translate_text": "Что такое электрическое напряжение?",
                "phrase_code": "pl",
                "language_level": "B1",
                "phrase_locale": "pl-PL",
                "phrase_voice": "pl-PL-MarekNeural",
                "phrase_voice_gender": "Male",
                "translate_code": "ru",
                "translate_locale": "ru-RU",
                "translate_voice": "ru-RU-SvetlanaNeural",
                "translate_voice_gender": "Female",
                "pause_ms": 1960,
                "speed": 1.0,
                "repeat_count": 1,
            },
            {
                "item_id": 39,
                "item_order": 2,
                "phrase_id": 58,
                "difficulty": 2,
                "phrase_text": "Jak się masz?",
                "translate_text": "Как дела?",
                "phrase_code": "pl",
                "language_level": "A2",
                "phrase_locale": "pl-PL",
                "phrase_voice": "pl-PL-MarekNeural",
                "phrase_voice_gender": "Male",
                "translate_code": "ru",
                "translate_locale": "ru-RU",
                "translate_voice": "ru-RU-SvetlanaNeural",
                "translate_voice_gender": "Female",
                "pause_ms": 1500,
                "speed": 1.0,
                "repeat_count": 2,
            },
        ],

        "state": {
            "current_index": 1,
        },
    })

    return session


def test_export_creates_expected_sheets(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)

    assert workbook.sheetnames == [
        "Set",
        "Items",
        "Voices",
    ]


def test_export_creates_expected_sheets(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)

    assert workbook.sheetnames == [
        "Set",
        "Items",
        "Voices",
    ]


def test_export_set_metadata(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Set"]

    assert worksheet["A1"].value == "Field"
    assert worksheet["B1"].value == "Value"

    assert worksheet["A2"].value == "set_index"
    assert worksheet["B2"].value == 2

    assert worksheet["A3"].value == "set_name"
    assert worksheet["B3"].value == "Polski B2"

    assert worksheet["A4"].value == "set_description"
    assert worksheet["B4"].value == "Testowy zestaw"


def test_export_items_header(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    assert headers == SessionExcel.ITEM_COLUMNS


def test_export_items_count(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    # Первая строка — заголовок.
    assert worksheet.max_row == len(session) + 1


def test_export_first_item(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    assert worksheet["A2"].value == 1
    assert worksheet["B2"].value == "pl"
    assert worksheet["C2"].value == "B1"

    assert worksheet["D2"].value == (
        "Co to jest napięcie elektryczne?"
    )

    assert worksheet["E2"].value == "ru"

    assert worksheet["F2"].value == (
        "Что такое электрическое напряжение?"
    )


def test_export_second_item(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    assert worksheet["A3"].value == 2
    assert worksheet["C3"].value == "A2"
    assert worksheet["D3"].value == "Jak się masz?"
    assert worksheet["F3"].value == "Как дела?"


def test_export_numeric_values(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    # Первый элемент
    assert worksheet["G2"].value == 1
    assert worksheet["H2"].value == 1960
    assert worksheet["I2"].value == 1.0
    assert worksheet["J2"].value == 1

    # Второй элемент
    assert worksheet["G3"].value == 2
    assert worksheet["H3"].value == 1500
    assert worksheet["I3"].value == 1.0
    assert worksheet["J3"].value == 2


def test_export_unicode(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Items"]

    assert worksheet["D2"].value == (
        "Co to jest napięcie elektryczne?"
    )

    assert worksheet["F2"].value == (
        "Что такое электрическое напряжение?"
    )


def test_export_does_not_contain_database_fields(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)

    set_sheet = workbook["Set"]
    items_sheet = workbook["Items"]

    set_fields = [
        set_sheet.cell(
            row=row,
            column=1,
        ).value
        for row in range(
            2,
            set_sheet.max_row + 1,
        )
    ]

    item_headers = [
        cell.value
        for cell in items_sheet[1]
    ]

    assert "set_id" not in set_fields
    assert "user_id" not in set_fields
    assert "user_nickname" not in set_fields

    assert "item_id" not in item_headers
    assert "phrase_id" not in item_headers
    assert "set_id" not in item_headers
    assert "user_id" not in item_headers


def test_export_does_not_contain_session_state(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
    )

    workbook = load_workbook(filename)

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                assert cell.value != "current_index"
def create_test_voices() -> list[dict]:
    """
    Создаёт небольшой тестовый справочник голосов.

    Реальный Edge TTS в unit-тестах не используется.
    """

    return [
        {
            "short_name": "pl-PL-MarekNeural",
            "locale": "pl-PL",
            "locale_name": "Polish (Poland)",
            "gender": "Male",
            "friendly_name": "Marek",
        },
        {
            "short_name": "pl-PL-ZofiaNeural",
            "locale": "pl-PL",
            "locale_name": "Polish (Poland)",
            "gender": "Female",
            "friendly_name": "Zofia",
        },
        {
            "short_name": "ru-RU-SvetlanaNeural",
            "locale": "ru-RU",
            "locale_name": "Russian (Russia)",
            "gender": "Female",
            "friendly_name": "Svetlana",
        },
    ]


def test_export_voices_header(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
        voices=create_test_voices(),
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Voices"]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    assert headers == SessionExcel.VOICE_COLUMNS


def test_export_voices_count(tmp_path):
    session = create_test_session()

    voices = create_test_voices()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
        voices=voices,
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Voices"]

    assert worksheet.max_row == len(voices) + 1


def test_export_first_voice(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
        voices=create_test_voices(),
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Voices"]

    assert worksheet["A2"].value == "pl"
    assert worksheet["B2"].value == "pl-PL"
    assert worksheet["C2"].value == "Polish (Poland)"
    assert worksheet["D2"].value == "pl-PL-MarekNeural"
    assert worksheet["E2"].value == "Male"
    assert worksheet["F2"].value == "Marek"


def test_export_second_voice(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
        voices=create_test_voices(),
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Voices"]

    assert worksheet["A3"].value == "pl"
    assert worksheet["B3"].value == "pl-PL"
    assert worksheet["D3"].value == "pl-PL-ZofiaNeural"
    assert worksheet["E3"].value == "Female"


def test_export_voices_empty(tmp_path):
    session = create_test_session()

    filename = tmp_path / "test_set.xlsx"

    SessionExcel.export(
        session,
        filename,
        voices=[],
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Voices"]

    assert worksheet.max_row == 1

    assert [
        cell.value
        for cell in worksheet[1]
    ] == SessionExcel.VOICE_COLUMNS