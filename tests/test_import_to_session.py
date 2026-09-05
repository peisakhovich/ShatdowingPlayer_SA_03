from session.session_excel import SessionExcel

def test_import_to_session(tmp_path):
    from openpyxl import Workbook

    excel_file = tmp_path / "test_import_to_session.xlsx"

    # ---------------------------------------------------------
    # Создаём Excel с новым набором
    # ---------------------------------------------------------

    workbook = Workbook()

    set_sheet = workbook.active
    set_sheet.title = "Set"

    set_sheet.append(["Field", "Value"])
    set_sheet.append(["set_index", 25])
    set_sheet.append(["set_name", "Imported Set"])
    set_sheet.append([
        "set_description",
        "Imported description",
    ])

    items_sheet = workbook.create_sheet("Items")
    items_sheet.append(SessionExcel.ITEM_COLUMNS)

    items_sheet.append([
        1,
        "pl",
        "B1",
        "Pierwsze zdanie",
        "ru",
        "Первое предложение",
        2,
        1500,
        1.0,
        2,
        "pl-PL",
        "pl-PL-MarekNeural",
        "Male",
        "ru-RU",
        "ru-RU-SvetlanaNeural",
        "Female",
    ])

    items_sheet.append([
        5,
        "pl",
        "B2",
        "Drugie zdanie",
        "ru",
        "Второе предложение",
        3,
        2000,
        0.9,
        1,
        "pl-PL",
        "pl-PL-ZofiaNeural",
        "Female",
        "ru-RU",
        "ru-RU-DmitryNeural",
        "Male",
    ])

    workbook.create_sheet("Voices")

    workbook.save(excel_file)
    workbook.close()

    # ---------------------------------------------------------
    # Создаём существующую Session
    # ---------------------------------------------------------

    from session.session import Session

    session = Session()

    session.set_user(
        user_id=2,
        user_nickname="pga",
    )

    session.load_data({
        "user_id": 2,
        "user_nickname": "pga",
        "set": {
            "set_id": 999,
            "user_id": 2,
            "user_nickname": "pga",
            "set_index": 999,
            "set_name": "Old Set",
            "set_description": "Old description",
        },
        "items": [
            {
                "item_order": 99,
                "phrase_code": "old",
                "phrase_text": "Old phrase",
            }
        ],
        "state": {
            "current_index": 10,
        },
    })

    # ---------------------------------------------------------
    # Выполняем Excel -> Session
    # ---------------------------------------------------------

    result = SessionExcel.import_(
        session,
        excel_file,
    )

    # Метод должен вернуть ту же Session
    assert result is session

    # ---------------------------------------------------------
    # Проверяем пользователя
    # ---------------------------------------------------------

    assert session.user_id == 2
    assert session.user_nickname == "pga"

    # ---------------------------------------------------------
    # Проверяем, что старый Set полностью заменён
    # ---------------------------------------------------------

    assert session._set == {
        "set_index": 25,
        "set_name": "Imported Set",
        "set_description": "Imported description",
    }

    # ---------------------------------------------------------
    # Проверяем, что старые items полностью заменены
    # ---------------------------------------------------------

    assert session._items == [
        {
            "item_order": 1,
            "phrase_code": "pl",
            "language_level": "B1",
            "phrase_text": "Pierwsze zdanie",
            "translate_code": "ru",
            "translate_text": "Первое предложение",
            "difficulty": 2,
            "pause_ms": 1500,
            "speed": 1.0,
            "repeat_count": 2,
            "phrase_locale": "pl-PL",
            "phrase_voice": "pl-PL-MarekNeural",
            "phrase_voice_gender": "Male",
            "translate_locale": "ru-RU",
            "translate_voice": "ru-RU-SvetlanaNeural",
            "translate_voice_gender": "Female",
        },
        {
            "item_order": 5,
            "phrase_code": "pl",
            "language_level": "B2",
            "phrase_text": "Drugie zdanie",
            "translate_code": "ru",
            "translate_text": "Второе предложение",
            "difficulty": 3,
            "pause_ms": 2000,
            "speed": 0.9,
            "repeat_count": 1,
            "phrase_locale": "pl-PL",
            "phrase_voice": "pl-PL-ZofiaNeural",
            "phrase_voice_gender": "Female",
            "translate_locale": "ru-RU",
            "translate_voice": "ru-RU-DmitryNeural",
            "translate_voice_gender": "Male",
        },
    ]

    # item_order не перенумеровывается
    assert session._items[1]["item_order"] == 5

    # ---------------------------------------------------------
    # После импорта позиция воспроизведения начинается с 0
    # ---------------------------------------------------------

    assert session.current_index == 0

def test_import_real_excel_to_session():
    from session.session import Session

    excel_file = "test_output/session_export.xlsx"

    session = Session()

    session.set_user(
        user_id=2,
        user_nickname="pga",
    )

    session.load_data({
        "user_id": 2,
        "user_nickname": "pga",
        "set": {
            "set_id": 999,
            "set_name": "Old Set",
            "set_description": "Old description",
        },
        "items": [
            {
                "item_order": 999,
                "phrase_code": "old",
                "phrase_text": "Old phrase",
            }
        ],
        "state": {
            "current_index": 10,
        },
    })

    result = SessionExcel.import_(
        session,
        excel_file,
    )

    assert result is session

    # Пользователь сохраняется
    assert session.user_id == 2
    assert session.user_nickname == "pga"

    # Set пришёл из Excel
    assert session._set["set_index"] == 2
    assert session._set["set_name"] == "Polisн  B2"
    assert (
        session._set["set_description"]
        == "Różne rozmowy na tematy biznesowe"
    )

    # В текущем Excel 14 строк Items
    assert len(session._items) == 14

    # Проверяем, что импорт действительно заменил старые данные
    assert session._items[0]["item_order"] == 1
    assert session._items[-1]["item_order"] == 26

    # Проверяем реальные данные
    assert (
        session._items[0]["phrase_text"]
        == "Co to jest napięcie elektryczne?"
    )

    assert (
        session._items[0]["translate_text"]
        == "Что такое электрическое напряжение?"
    )

    # Позиция воспроизведения после импорта сброшена
    assert session.current_index == 0