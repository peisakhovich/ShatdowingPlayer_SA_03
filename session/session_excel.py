from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


class SessionExcel:
    """
    Обмен данными Session с Excel.

    Excel-файл содержит три видимых листа:

        Set
            параметры Training Set

        Items
            элементы Training Set

        Voices
            справочник голосов Edge TTS

    На данном этапе реализован:
        Session -> Excel
    """

    SET_SHEET = "Set"
    ITEMS_SHEET = "Items"
    VOICES_SHEET = "Voices"

    SET_FIELDS = [
        "set_index",
        "set_name",
        "set_description",
    ]

    ITEM_COLUMNS = [
        "item_order",
        "phrase_code",
        "language_level",
        "phrase_text",
        "translate_code",
        "translate_text",
        "difficulty",
        "pause_ms",
        "speed",
        "repeat_count",
        "phrase_locale",
        "phrase_voice",
        "phrase_voice_gender",
        "translate_locale",
        "translate_voice",
        "translate_voice_gender",
    ]

    VOICE_COLUMNS = [
        "code",
        "locale",
        "locale_name",
        "short_name",
        "gender",
        "friendly_name",
    ]

    LANGUAGE_LEVELS = [
        "A1",
        "A2",
        "B1",
        "B2",
        "C1",
        "C2",
    ]

    GENDERS = [
        "Male",
        "Female",
    ]

    DIFFICULTIES = [
        1,
        2,
        3,
        4,
        5,
    ]

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    @classmethod
    def export(
        cls,
        session,
        filename: str | Path,
        voices: list[dict] | None = None,
    ) -> Path:
        """
        Экспортирует текущую Session в Excel.

        Parameters
        ----------
        session:
            Экземпляр Session.

        filename:
            Путь к создаваемому xlsx-файлу.

        voices:
            Список голосов Edge TTS.

            Передаётся готовым списком специально для того,
            чтобы SessionExcel не зависел непосредственно
            от TTS и edge_tts.

        Returns
        -------
        Path
            Путь к созданному Excel-файлу.
        """

        filename = Path(filename)

        data = session.get_data()

        set_data = data.get("set", {})
        items = data.get("items", [])

        workbook = Workbook()

        set_sheet = workbook.active
        set_sheet.title = cls.SET_SHEET

        items_sheet = workbook.create_sheet(
            cls.ITEMS_SHEET
        )

        voices_sheet = workbook.create_sheet(
            cls.VOICES_SHEET
        )

        cls._write_set_sheet(
            worksheet=set_sheet,
            set_data=set_data,
        )

        cls._write_items_sheet(
            worksheet=items_sheet,
            items=items,
        )

        cls._write_voices_sheet(
            worksheet=voices_sheet,
            voices=voices or [],
        )

        # Создаём списки выбора для Items.
        cls._add_items_validations(
            worksheet=items_sheet,
            voices_worksheet=voices_sheet,
            max_rows=max(len(items) + 100, 100),
        )

        workbook.save(filename)

        return filename

    # ------------------------------------------------------------------
    # Set
    # ------------------------------------------------------------------

    @classmethod
    def _write_set_sheet(
        cls,
        worksheet,
        set_data: dict,
    ) -> None:
        """
        Записывает информацию о Training Set
        на лист Set.
        """

        worksheet.append(["Field", "Value"])

        for field in cls.SET_FIELDS:
            worksheet.append([
                field,
                set_data.get(field, ""),
            ])

        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 60

    # ------------------------------------------------------------------
    # Items
    # ------------------------------------------------------------------

    @classmethod
    def _write_items_sheet(
        cls,
        worksheet,
        items: list[dict],
    ) -> None:
        """
        Записывает элементы Training Set
        на лист Items.
        """

        worksheet.append(cls.ITEM_COLUMNS)

        for item in items:

            row = [
                item.get(column, "")
                for column in cls.ITEM_COLUMNS
            ]

            worksheet.append(row)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        cls._set_column_widths(worksheet)

    # ------------------------------------------------------------------
    # Voices
    # ------------------------------------------------------------------

    @classmethod
    def _write_voices_sheet(
        cls,
        worksheet,
        voices: list[dict],
    ) -> None:
        """
        Записывает справочник голосов Edge TTS
        на лист Voices.

        В колонках A:F находится видимый справочник.

        Начиная с H создаются скрытые служебные списки,
        используемые Excel Data Validation.
        """

        worksheet.append(cls.VOICE_COLUMNS)

        for voice in voices:

            locale = voice.get("locale", "")

            code = ""

            if locale:
                code = locale.split("-")[0].lower()

            worksheet.append([
                code,
                locale,
                voice.get("locale_name", ""),
                voice.get("short_name", ""),
                voice.get("gender", ""),
                voice.get("friendly_name", ""),
            ])

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        cls._set_column_widths(worksheet)

        # Служебные списки для Data Validation.
        cls._write_validation_lists(
            worksheet=worksheet,
            voices=voices,
        )

    # ------------------------------------------------------------------
    # Validation lists
    # ------------------------------------------------------------------

    @classmethod
    def _write_validation_lists(
        cls,
        worksheet,
        voices: list[dict],
    ) -> None:
        """
        Создаёт скрытые служебные списки на листе Voices.

        Эти списки используются зависимыми выпадающими
        списками на листе Items.
        """

        # Уникальные языки.
        languages = sorted({
            voice.get("locale", "").split("-")[0].lower()
            for voice in voices
            if voice.get("locale")
        })

        # Уникальные locale для каждого языка.
        locales_by_language: dict[str, set[str]] = {}

        # Уникальные голоса для каждого locale.
        voices_by_locale: dict[str, set[str]] = {}

        for voice in voices:

            locale = voice.get("locale", "")
            short_name = voice.get("short_name", "")

            if not locale:
                continue

            language = locale.split("-")[0].lower()

            locales_by_language.setdefault(
                language,
                set(),
            ).add(locale)

            if short_name:
                voices_by_locale.setdefault(
                    locale,
                    set(),
                ).add(short_name)

        # Начинаем со столбца H.
        column = 8

        # --------------------------------------------------------------
        # Языки
        # --------------------------------------------------------------

        worksheet.cell(
            row=1,
            column=column,
            value="languages",
        )

        for row, language in enumerate(
            languages,
            start=2,
        ):
            worksheet.cell(
                row=row,
                column=column,
                value=language,
            )

        language_column = worksheet.cell(
            row=1,
            column=column,
        ).column_letter

        cls._add_defined_name(
            name="language_codes",
            worksheet=worksheet,
            column_letter=language_column,
            start_row=2,
            end_row=max(1, len(languages) + 1),
        )

        column += 1

        # --------------------------------------------------------------
        # Language levels
        # --------------------------------------------------------------

        worksheet.cell(
            row=1,
            column=column,
            value="language_levels",
        )

        for row, level in enumerate(
            cls.LANGUAGE_LEVELS,
            start=2,
        ):
            worksheet.cell(
                row=row,
                column=column,
                value=level,
            )

        level_column = worksheet.cell(
            row=1,
            column=column,
        ).column_letter

        cls._add_defined_name(
            name="language_levels",
            worksheet=worksheet,
            column_letter=level_column,
            start_row=2,
            end_row=len(cls.LANGUAGE_LEVELS) + 1,
        )

        column += 1

        # --------------------------------------------------------------
        # Gender
        # --------------------------------------------------------------

        worksheet.cell(
            row=1,
            column=column,
            value="genders",
        )

        for row, gender in enumerate(
            cls.GENDERS,
            start=2,
        ):
            worksheet.cell(
                row=row,
                column=column,
                value=gender,
            )

        gender_column = worksheet.cell(
            row=1,
            column=column,
        ).column_letter

        cls._add_defined_name(
            name="voice_genders",
            worksheet=worksheet,
            column_letter=gender_column,
            start_row=2,
            end_row=len(cls.GENDERS) + 1,
        )

        column += 1

        # --------------------------------------------------------------
        # Difficulty
        # --------------------------------------------------------------

        worksheet.cell(
            row=1,
            column=column,
            value="difficulties",
        )

        for row, difficulty in enumerate(
            cls.DIFFICULTIES,
            start=2,
        ):
            worksheet.cell(
                row=row,
                column=column,
                value=difficulty,
            )

        difficulty_column = worksheet.cell(
            row=1,
            column=column,
        ).column_letter

        cls._add_defined_name(
            name="difficulties",
            worksheet=worksheet,
            column_letter=difficulty_column,
            start_row=2,
            end_row=len(cls.DIFFICULTIES) + 1,
        )

        column += 1

        # --------------------------------------------------------------
        # Locales by language
        # --------------------------------------------------------------

        for language in languages:

            values = sorted(
                locales_by_language.get(
                    language,
                    set(),
                )
            )

            if not values:
                continue

            worksheet.cell(
                row=1,
                column=column,
                value=f"locales_{language}",
            )

            for row, value in enumerate(
                values,
                start=2,
            ):
                worksheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

            column_letter = worksheet.cell(
                row=1,
                column=column,
            ).column_letter

            cls._add_defined_name(
                name=f"locales_{language}",
                worksheet=worksheet,
                column_letter=column_letter,
                start_row=2,
                end_row=len(values) + 1,
            )

            column += 1

        # --------------------------------------------------------------
        # Voices by locale
        # --------------------------------------------------------------

        for locale in sorted(voices_by_locale):

            values = sorted(
                voices_by_locale[locale]
            )

            if not values:
                continue

            safe_locale = locale.replace(
                "-",
                "_",
            )

            worksheet.cell(
                row=1,
                column=column,
                value=f"voices_{safe_locale}",
            )

            for row, value in enumerate(
                values,
                start=2,
            ):
                worksheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

            column_letter = worksheet.cell(
                row=1,
                column=column,
            ).column_letter

            cls._add_defined_name(
                name=f"voices_{safe_locale}",
                worksheet=worksheet,
                column_letter=column_letter,
                start_row=2,
                end_row=len(values) + 1,
            )

            column += 1

        # Скрываем служебные колонки H и далее.
        for index in range(
            8,
            column,
        ):
            worksheet.column_dimensions[
                worksheet.cell(
                    row=1,
                    column=index,
                ).column_letter
            ].hidden = True

    # ------------------------------------------------------------------
    # Data Validation
    # ------------------------------------------------------------------

    @classmethod
    def _add_items_validations(
        cls,
        worksheet,
        voices_worksheet,
        max_rows: int,
    ) -> None:
        """
        Добавляет выпадающие списки на лист Items.

        Колонки Items:

            B  phrase_code
            C  language_level
            E  translate_code
            G  difficulty
            K  phrase_locale
            L  phrase_voice
            M  phrase_voice_gender
            N  translate_locale
            O  translate_voice
            P  translate_voice_gender
        """

        # --------------------------------------------------------------
        # phrase_code
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=language_codes",
            allow_blank=True,
        )

        validation.error = "Выберите язык из списка."
        validation.errorTitle = "Неверный язык"
        validation.prompt = "Выберите код языка."
        validation.promptTitle = "phrase_code"

        worksheet.add_data_validation(validation)

        validation.add(
            f"B2:B{max_rows}"
        )

        # --------------------------------------------------------------
        # language_level
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=language_levels",
            allow_blank=True,
        )

        validation.error = "Выберите уровень A1-C2."
        validation.errorTitle = "Неверный уровень"

        worksheet.add_data_validation(validation)

        validation.add(
            f"C2:C{max_rows}"
        )

        # --------------------------------------------------------------
        # translate_code
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=language_codes",
            allow_blank=True,
        )

        validation.error = "Выберите язык из списка."
        validation.errorTitle = "Неверный язык"

        worksheet.add_data_validation(validation)

        validation.add(
            f"E2:E{max_rows}"
        )

        # --------------------------------------------------------------
        # difficulty
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=difficulties",
            allow_blank=True,
        )

        validation.error = "Выберите значение от 1 до 5."
        validation.errorTitle = "Неверная сложность"

        worksheet.add_data_validation(validation)

        validation.add(
            f"G2:G{max_rows}"
        )

        # --------------------------------------------------------------
        # phrase_locale
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1='=INDIRECT("locales_"&B2)',
            allow_blank=True,
        )

        validation.error = (
            "Сначала выберите phrase_code."
        )
        validation.errorTitle = "Неверный locale"

        worksheet.add_data_validation(validation)

        validation.add(
            f"K2:K{max_rows}"
        )

        # --------------------------------------------------------------
        # phrase_voice
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1='=INDIRECT("voices_"&SUBSTITUTE(K2,"-","_"))',
            allow_blank=True,
        )

        validation.error = (
            "Сначала выберите phrase_locale."
        )
        validation.errorTitle = "Неверный голос"

        worksheet.add_data_validation(validation)

        validation.add(
            f"L2:L{max_rows}"
        )

        # --------------------------------------------------------------
        # phrase_voice_gender
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=voice_genders",
            allow_blank=True,
        )

        validation.error = "Выберите Male или Female."
        validation.errorTitle = "Неверный пол голоса"

        worksheet.add_data_validation(validation)

        validation.add(
            f"M2:M{max_rows}"
        )

        # --------------------------------------------------------------
        # translate_locale
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1='=INDIRECT("locales_"&E2)',
            allow_blank=True,
        )

        validation.error = (
            "Сначала выберите translate_code."
        )
        validation.errorTitle = "Неверный locale"

        worksheet.add_data_validation(validation)

        validation.add(
            f"N2:N{max_rows}"
        )

        # --------------------------------------------------------------
        # translate_voice
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1='=INDIRECT("voices_"&SUBSTITUTE(N2,"-","_"))',
            allow_blank=True,
        )

        validation.error = (
            "Сначала выберите translate_locale."
        )
        validation.errorTitle = "Неверный голос"

        worksheet.add_data_validation(validation)

        validation.add(
            f"O2:O{max_rows}"
        )

        # --------------------------------------------------------------
        # translate_voice_gender
        # --------------------------------------------------------------

        validation = DataValidation(
            type="list",
            formula1="=voice_genders",
            allow_blank=True,
        )

        validation.error = "Выберите Male или Female."
        validation.errorTitle = "Неверный пол голоса"

        worksheet.add_data_validation(validation)

        validation.add(
            f"P2:P{max_rows}"
        )

    # ------------------------------------------------------------------
    # Defined names
    # ------------------------------------------------------------------

    @staticmethod
    def _add_defined_name(
        name: str,
        worksheet,
        column_letter: str,
        start_row: int,
        end_row: int,
    ) -> None:
        """
        Создаёт именованный диапазон Excel.
        """

        reference = (
            f"'{worksheet.title}'!"
            f"${column_letter}${start_row}:"
            f"${column_letter}${end_row}"
        )

        defined_name = DefinedName(
            name,
            attr_text=reference,
        )

        worksheet.parent.defined_names.add(
            defined_name
        )

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    @staticmethod
    def _set_column_widths(
        worksheet,
    ) -> None:
        """
        Устанавливает удобную ширину колонок.
        """

        for column_cells in worksheet.columns:

            column_letter = column_cells[0].column_letter

            max_length = 0

            for cell in column_cells:

                value = cell.value

                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                50,
            )